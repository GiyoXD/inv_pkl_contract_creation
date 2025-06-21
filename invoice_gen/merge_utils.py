import openpyxl
import traceback
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
# from openpyxl.worksheet.dimensions import RowDimension # Not strictly needed for access
from typing import Dict, List, Optional, Tuple, Any

# --- store_original_merges FILTERED to ignore merges ABOVE row 16 ---
def store_original_merges(workbook: openpyxl.Workbook, sheet_names: List[str]) -> Dict[str, List[Tuple[int, Any, Optional[float]]]]:
    """
    Stores the HORIZONTAL span (colspan), the value of the top-left cell,
    and the height of the starting row for merged ranges in specified sheets,
    ASSUMING all merges are only 1 row high AND **start at row 16 or below**.
    Merges starting above row 16 (row < 16) are ignored.
    WARNING: Does NOT store starting coordinates... (rest of docstring unchanged)

    Args: (args unchanged)

    Returns:
        A dictionary where keys are sheet names and values are lists of
        tuples: (col_span, top_left_cell_value, row_height).
        row_height will be None if the original row had default height.
    """
    original_merges = {}
    print("\nStoring original merge horizontal spans, top-left values, and row heights (NO coordinates)...")
    print("  (Ignoring merges that start above row 16)") # Updated filter info
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            worksheet: Worksheet = workbook[sheet_name] # Type hint for clarity
            merges_data = []
            merged_ranges_copy = list(worksheet.merged_cells.ranges)
            skipped_above_16_count = 0 # Counter for this filter

            for merged_range in merged_ranges_copy:
                min_col, min_row, max_col, max_row = merged_range.bounds

                # --- Check 1: Skip if multi-row ---
                if max_row != min_row:
                    # print(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - it spans multiple rows ({min_row} to {max_row}).")
                    continue

                # ***** NEW CHECK 2: Skip if merge starts ABOVE row 16 *****
                if min_row < 16:
                    # print(f"  Skipping merge {merged_range.coord} on sheet '{sheet_name}' - starts at row {min_row} (above row 16).") # Keep commented unless needed
                    skipped_above_16_count += 1
                    continue
                # ***** END NEW CHECK *****

                # --- If not skipped, proceed to get span, height, value ---
                col_span = max_col - min_col + 1
                row_height = None # Default to None
                try:
                    # Get Row Height
                    row_dim = worksheet.row_dimensions[min_row]
                    row_height = row_dim.height
                    # print(f"    DEBUG Store: Sheet='{sheet_name}', MergeCoord='{merged_range.coord}', StartRow={min_row}, Storing Height={row_height} (Type: {type(row_height)})")

                    # Get Value
                    top_left_value = worksheet.cell(row=min_row, column=min_col).value

                    # Store Data (span, value, height)
                    merges_data.append((col_span, top_left_value, row_height))

                except KeyError:
                     print(f"    Warning: Could not find row dimension for row {min_row} on sheet '{sheet_name}' while getting height. Storing height as None.")
                     try:
                         top_left_value = worksheet.cell(row=min_row, column=min_col).value
                     except Exception as val_e:
                         print(f"    Warning: Also failed to get value for merge at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value as None. Error: {val_e}")
                         top_left_value = None
                     merges_data.append((col_span, top_left_value, None))

                except Exception as e:
                    print(f"    Warning: Could not get value/height for merge starting at ({min_row},{min_col}) on sheet '{sheet_name}'. Storing value/height as None. Error: {e}")
                    merges_data.append((col_span, None, None))

            original_merges[sheet_name] = merges_data
            print(f"  Stored {len(original_merges[sheet_name])} horizontal merge span/value/height entries for sheet '{sheet_name}'.")
            # Report skipped count for this filter
            if skipped_above_16_count > 0:
                print(f"    (Skipped {skipped_above_16_count} merges starting above row 16)")
        else:
             print(f"  Warning: Sheet '{sheet_name}' specified but not found during merge storage.")
             original_merges[sheet_name] = []
    return original_merges

# --- find_and_restore_merges_heuristic remains unchanged (still searches bottom-up, applies stored value/height) ---
def find_and_restore_merges_heuristic(workbook: openpyxl.Workbook,
                                      stored_merges: Dict[str, List[Tuple[int, Any, Optional[float]]]],
                                      processed_sheet_names: List[str],
                                      search_range_str: str = "A16:H200"):
    """
    Attempts to restore merges based on stored HORIZONTAL spans, values, and row heights
    by searching for the value within a specified range (default A16:H200).
    This version is silent, with no detailed logging.

    WARNING: This is a HEURISTIC approach... (rest of docstring unchanged)

    Args: (args unchanged)
    """
    print("Starting merge restoration process...")

    # These counters are still used by the logic but are no longer printed.
    restored_count = 0
    failed_count = 0
    skipped_count = 0
    skipped_duplicate_value_count = 0

    # --- Define search boundaries (critical errors are still reported) ---
    try:
        search_min_col, search_min_row, search_max_col, search_max_row = range_boundaries(search_range_str)
    except TypeError as te:
        print(f"Error processing search range '{search_range_str}'. Check openpyxl version compatibility or range format. Internal error: {te}")
        traceback.print_exc()
        return
    except Exception as e:
        print(f"Error: Invalid search range string '{search_range_str}'. Cannot proceed with restoration. Error: {e}")
        return

    # --- Loop through sheets ---
    for sheet_name in processed_sheet_names:
        if sheet_name in workbook.sheetnames and sheet_name in stored_merges:
            worksheet: Worksheet = workbook[sheet_name]
            original_merges_data = stored_merges[sheet_name]
            successfully_restored_values_on_sheet = set()

            # --- Loop through stored merge info ---
            for col_span, stored_value, stored_height in original_merges_data:

                if col_span <= 1:
                    skipped_count += 1
                    continue

                if stored_value in successfully_restored_values_on_sheet:
                    skipped_duplicate_value_count += 1
                    continue

                found = False
                # --- Search range loop - ROW SEARCH REVERSED ---
                for r in range(search_max_row, search_min_row - 1, -1):
                    for c in range(search_min_col, search_max_col + 1):
                        current_cell = worksheet.cell(row=r, column=c)
                        current_val = current_cell.value

                        if current_val == stored_value:
                            start_row, start_col = r, c
                            end_row = start_row
                            end_col = start_col + col_span - 1

                            # --- Proactively unmerge any conflicting ranges ---
                            merged_ranges_copy = list(worksheet.merged_cells.ranges)
                            for existing_merge in merged_ranges_copy:
                                rows_overlap = (existing_merge.min_row <= end_row) and (existing_merge.max_row >= start_row)
                                cols_overlap = (existing_merge.min_col <= end_col) and (existing_merge.max_col >= start_col)

                                if rows_overlap and cols_overlap:
                                    try:
                                        worksheet.unmerge_cells(str(existing_merge))
                                    except Exception:
                                        # Fails silently as requested
                                        pass

                            # --- Apply the new merge, Row Height, AND Value ---
                            try:
                                worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

                                if stored_height is not None:
                                    try:
                                        worksheet.row_dimensions[start_row].height = stored_height
                                    except Exception:
                                        # Fails silently
                                        pass

                                top_left_cell_to_set = worksheet.cell(row=start_row, column=start_col)
                                top_left_cell_to_set.value = stored_value

                                successfully_restored_values_on_sheet.add(stored_value)
                                restored_count += 1
                                found = True
                                break

                            except Exception:
                                failed_count += 1
                                found = True
                                break

                    if found:
                        break

                if not found:
                    if stored_value not in successfully_restored_values_on_sheet:
                        failed_count += 1

    print("Merge restoration process finished.")