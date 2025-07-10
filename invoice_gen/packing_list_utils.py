import invoice_utils
import style_utils
import merge_utils
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple

def calculate_rows_to_generate(packing_list_data: dict, sheet_config: dict) -> int:
    """
    Calculates the total number of rows required to generate the full packing list.

    This function determines the necessary space by summing the rows needed for headers,
    data, footers for each table, plus spacing and any grand totals.

    Args:
        packing_list_data: The core data object containing raw_data.
        sheet_config: The configuration for the sheet.

    Returns:
        The total integer count of rows that will be written.
    """
    raw_data = packing_list_data.get('raw_data', {})
    if not raw_data:
        return 0

    num_tables = len(raw_data)
    header_rows_per_table = 2
    footer_rows_per_table = 1
    spacing_rows = max(0, num_tables - 1)
    grand_total_rows = 1 if num_tables > 1 else 0
    total_data_rows = sum(len(table_data.get('net', [])) for table_data in raw_data.values())

    total_generated_rows = (
        (header_rows_per_table * num_tables) +
        total_data_rows +
        (footer_rows_per_table * num_tables) +
        spacing_rows +
        grand_total_rows
    )
    print(f"  Calculated that {total_generated_rows} rows will be generated.")
    return total_generated_rows

def generate_full_packing_list(worksheet: Worksheet, start_row: int, packing_list_data: dict, sheet_config: dict):
    """
    Generates the entire packing list content, including headers, data, and footers.
    """
    write_pointer_row = start_row
    all_data_ranges: List[Tuple[int, int]] = []
    all_header_infos: List[Dict] = []
    all_footer_rows: List[int] = []
    grand_total_pallets = 0

    header_to_write = sheet_config.get("header_to_write", [])
    footer_config = sheet_config.get("footer_configurations", {})
    styling_config = sheet_config.get("styling", {})
    mappings = sheet_config.get("mappings", {})
    data_map = mappings.get("data_map", {})
    static_col_values = mappings.get("initial_static", {}).get("values", [])
    
    raw_data = packing_list_data.get('raw_data', {})
    table_keys = sorted(raw_data.keys())
    num_tables = len(table_keys)

    for i, table_key in enumerate(table_keys):
        table_data = raw_data[table_key]
        num_data_rows = len(table_data.get('net', []))
        
        # 1. WRITE HEADER
        header_info = invoice_utils.write_header(worksheet, write_pointer_row, header_to_write, styling_config)
        all_header_infos.append(header_info)
        write_pointer_row = header_info.get('second_row_index', write_pointer_row) + 1
        col_map = header_info.get('column_id_map', {})
        num_columns = header_info.get('num_columns', 1)
        static_col_idx = col_map.get(mappings.get("initial_static", {}).get("column_header_id"))
        idx_to_id_map = {v: k for k, v in col_map.items()}

        # 2. WRITE DATA AND APPLY STYLING
        data_start_row = write_pointer_row
        for r_idx in range(num_data_rows):
            current_row = write_pointer_row + r_idx
            if static_col_idx and r_idx < len(static_col_values):
                worksheet.cell(row=current_row, column=static_col_idx).value = static_col_values[r_idx]
            for data_key, mapping_info in data_map.items():
                if col_idx := col_map.get(mapping_info.get("id")):
                    if data_key in table_data:
                        worksheet.cell(row=current_row, column=col_idx).value = table_data[data_key][r_idx]
            for c_idx in range(1, num_columns + 1):
                cell = worksheet.cell(row=current_row, column=c_idx)
                style_context = {
                    "col_id": idx_to_id_map.get(c_idx), "col_idx": c_idx,
                    "static_col_idx": static_col_idx, "row_index": r_idx,
                    "num_data_rows": num_data_rows
                }
                style_utils.apply_cell_style(cell, styling_config, style_context)
        
        write_pointer_row += num_data_rows
        all_data_ranges.append((data_start_row, write_pointer_row - 1))

        # 2A. APPLY VERTICAL MERGES (THIS BLOCK WAS MISSING)
        data_end_row = write_pointer_row - 1
        vertical_merge_ids = mappings.get("vertical_merge_on_id", [])
        if vertical_merge_ids:
            print(f"Applying vertical merges for table '{table_key}'...")
            for col_id_to_merge in vertical_merge_ids:
                if col_idx := col_map.get(col_id_to_merge):
                    merge_utils.merge_vertical_cells_in_range(
                        worksheet=worksheet,
                        scan_col=col_idx,
                        start_row=data_start_row,
                        end_row=data_end_row
                    )

        # 3. WRITE FOOTER
        pallet_count = len(table_data.get('pallet_count', []))
        grand_total_pallets += pallet_count
        invoice_utils.write_footer_row(worksheet, write_pointer_row, header_info, [(data_start_row, write_pointer_row - 1)], footer_config, pallet_count)
        footer_merge_rules = footer_config.get("footer_merge_rules")
        merge_utils.apply_row_merges(worksheet, write_pointer_row, num_columns, footer_merge_rules)
        all_footer_rows.append(write_pointer_row)
        write_pointer_row += 1
        if i < num_tables - 1:
            write_pointer_row += 1

    # --- AFTER LOOP ---
    if num_tables > 1:
        last_header_info = all_header_infos[-1]
        num_columns = last_header_info.get('num_columns', 1)
        invoice_utils.write_footer_row(worksheet, write_pointer_row, last_header_info, all_data_ranges, footer_config, grand_total_pallets, "GRAND TOTAL:")
        grand_total_merge_rules = footer_config.get("grand_total_merge_rules")
        merge_utils.apply_row_merges(worksheet, write_pointer_row, num_columns, grand_total_merge_rules)
        all_footer_rows.append(write_pointer_row)

    # --- FINAL STYLING ---
    style_utils.apply_row_heights(worksheet, styling_config, all_header_infos, all_data_ranges, all_footer_rows)
    widths = styling_config.get("column_id_widths", {})
    col_map_text = {item['text']: item['id'] for item in header_to_write if 'text' in item and 'id' in item}
    for text, width in widths.items():
        if col_id := col_map_text.get(text):
            if col_idx := all_header_infos[0]['column_id_map'].get(col_id):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width